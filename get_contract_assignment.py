#!/usr/bin/env python3
import json
import logging
import sys
import os
import argparse
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class ACI:
	# __init__ is the constructor
	def __init__(self,fn):

		self.filename = fn
		logging.basicConfig(format='%(levelname)s:%(message)s', level=logging.DEBUG)

		if not os.path.isfile(self.filename):
			logging.error(" File %s not found" % filename)
			sys.exit(9)


	def load_file(self):
		with open(self.filename) as json_file:
			data = json.load(json_file)

		return data

	def set_col_width(self,ws):

		dims = {}
		for row in ws.rows:
			for cell in row:
				if cell.value:
					dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
		for col, value in dims.items():
			ws.column_dimensions[get_column_letter(col)].width = value

	def print_contract_data(self,data):
		contract_translation = {}
		contract_translation['fvRsProv'] = "Provided"
		contract_translation['fvRsCons'] = "Consumed"

		sub_type = {}
		sub_type['fvAp'] = 'AppProfile'
		sub_type['l3extOut'] = 'L3OUT'

		wb = Workbook()
		file = 'CONTRACT_ASSIGNMENTS.xlsx'
		ws = wb.active
		ws.title = 'contract_assignments'

		row = 1
		cell = ws.cell(column=1, row=row)
		cell.value = 'TENANT'

		cell = ws.cell(column=2, row=row)
		cell.value = 'INSTANCE_TYPE'

		cell = ws.cell(column=3, row=row)
		cell.value = 'INSTANCE_NAME'

		cell = ws.cell(column=4, row=row)
		cell.value = 'EPG'

		cell = ws.cell(column=5, row=row)
		cell.value = 'CONTRACT_TYPE'

		cell = ws.cell(column=6, row=row)
		cell.value = 'CONTRACT_NAME'

		row = 2

		for d in data['polUni']['children']:
			for keys in d.keys():
				if keys == 'fvTenant':
					tenant = d[keys]['attributes']['name']
					for children in d[keys]['children']:
						for c_keys in children.keys():
							if c_keys in ('fvAp','l3extOut'):
								sub_type_val = children[c_keys]['attributes']['name']
								for instance in children[c_keys]['children']:
									for instance_keys in instance.keys():
										if instance_keys in ('fvAEPg','l3extInstP'):
											epg = instance[instance_keys]['attributes']['name']
											for epg_children in instance[instance_keys]['children']:
												for epg_contract_keys in epg_children.keys():
													if epg_contract_keys == 'fvRsProv' or epg_contract_keys == 'fvRsCons':
														col = 1
														contract_name = epg_children[epg_contract_keys]['attributes']['tnVzBrCPName']
														#print(tenant,sub_type[c_keys],sub_type_val,epg,contract_translation[epg_contract_keys],contract_name, sep=',')
														cell = ws.cell(column=col, row=row)
														cell.value = tenant
														col+=1

														cell = ws.cell(column=col, row=row)
														cell.value = sub_type[c_keys]
														col+=1

														cell = ws.cell(column=col, row=row)
														cell.value = sub_type_val
														col+=1

														cell = ws.cell(column=col, row=row)
														cell.value = epg
														col+=1

														cell = ws.cell(column=col, row=row)
														cell.value = contract_translation[epg_contract_keys]
														col+=1

														cell = ws.cell(column=col, row=row)
														cell.value = contract_name

														row+=1

		self.set_col_width(ws)
		ws.sheet_view.zoomScale = 125

		# add filters
		ws.auto_filter.ref = 'A1:F900'

		wb.save(filename=file)


def read_arguments():
	parser = argparse.ArgumentParser(sys.argv[0] + " -f <ACI JSON File>")
	parser.add_argument("-f", "--input-file", dest="filename" , help="JSON From Backup File", required=True)
	args = parser.parse_args()
	return args

def main():
	args = read_arguments()
	data = ACI(args.filename)
	json_data = data.load_file()

	data.print_contract_data(json_data)
if __name__ == '__main__':
	main()